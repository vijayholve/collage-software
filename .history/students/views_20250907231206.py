from django.views.decorators.cache import never_cache
from django.contrib.auth.decorators import login_required
import os
from calendar import monthrange
from django.contrib.auth import login
from django.shortcuts import render, get_object_or_404, redirect 
from .models import Student, Attendance ,CustomUser ,Teacher ,hod ,ClassGroup ,Classes
from django.utils import timezone 
from django.template.defaulttags import register 
from base.forms import studuntform ,subjectform  
from django.db.models import Q  
from datetime import datetime, timedelta
from django.shortcuts import render
from .models import Attendance, Student, ClassGroup, Holiday
from .seed import teaches_to_send_mail ,student_to_send_mail
from django.contrib.auth.decorators import login_required
from django.conf import settings
from django.http import HttpResponse
from django.template.loader import render_to_string
from datetime import datetime, timedelta
from .task import send_mail_to_all_task_teacher ,send_mail_task , send_mail_to_all_task_students
from django import template
import pandas as pd
from django.http import FileResponse, Http404 ,HttpResponse 
from openpyxl import load_workbook
from faker import Faker
from django.contrib.messages import error
register = template.Library()

@register.filter(name='get_item')
def get_item(dictionary, key): 
    return dictionary.get(key)
from .form import holidays_form

def get_month_name(month_number):
    months = [
        "January", "February", "March", "April", "May", "June",
        "July", "August", "September", "October", "November", "December"
    ]
    
    if 1 <= month_number <= 12:
        return months[month_number - 1]
    else:
        return "Invalid month number"
def download_id_card(request, student_id):
    student = get_object_or_404(Student, pk=student_id)
    html_content = render_to_string('students/id_card.html', {'student': student})
    css_path = os.path.join(settings.BASE_DIR, 'static', 'students', 'id_card.css')
    
    with open(css_path, 'r') as css_file:
        css_content = css_file.read()
    
    # Inline the CSS content into the HTML content
    html_with_css = f"<style>{css_content}</style>{html_content}"
    
    if student.profile and student.profile.url.startswith('http'):
        student_image_url = student.profile.url  # External URL
    else:
        student_image_url = request.build_absolute_uri(student.profile.url) if student.profile else ''
    
    html_with_css = html_with_css.replace('{{ student_image }}', student_image_url)
    response = HttpResponse(html_with_css, content_type='text/html')
    response['Content-Disposition'] = f'attachment; filename="{student.name}_id_card.html"'
    return response

@login_required(login_url="login-page")
def home(request):
    classes=ClassGroup.objects.all() 
    user=request.user 
    user_type=user.user_type   
    print(user_type) 
    content={'username':user.username ,
             'user_type':user_type }
    return render(request,"students/home.html",content) 

# Create your views here.
@never_cache
@login_required(login_url='login-page')
def container(request):
    user = request.user
    type_u = user.user_type

    student = Student.objects.get(user=user)
    content = {
        'student': student,
    }
    return render(request, "students/containers.html", content)
        

@login_required(login_url="login-page")
def particular_student_attendance(request, sid): 
    student = Student.objects.get(id=sid) 
    today = datetime.now().date()
    start_date = today.replace(day=1)  
    end_date = today  
    holidays = Holiday.objects.filter(date__range=[start_date, end_date])
   
    holiday_days = {holiday.date.day: holiday.name for holiday in holidays}  
    

    if request.method == "POST":
        month_input = request.POST.get('month')
        if month_input:

            month = month_input.split('-')[1]
    else:
        month=today.month
        
    classgroup = student.classgroup
    attendance_data = {date: Attendance.objects.filter(student=student, date=date,present=True).exists()
                       for date in (start_date + timedelta(days=i) for i in range((end_date - start_date).days + 1))}
    dates = [start_date + timedelta(days=i) for i in range((end_date - start_date).days + 1)]
    first_day_of_month = start_date.weekday()
    total_days_in_month = monthrange(today.year, today.month)[1]
    calendar_days = [None] * first_day_of_month + dates + [None] * (6 - (first_day_of_month + total_days_in_month - 1) % 7)

    weeks = [calendar_days[i:i + 7] for i in range(0, len(calendar_days), 7)]
    month={i : get_month_name(i) for i in range(1,13) }
    holidays_count=holidays.count() 
    total_days = (end_date - start_date).days + 1
    sunday_count = sum(1 for i in range(total_days) 
                       if (start_date + timedelta(days=i)).weekday() == 6) 

    holidays_count += sunday_count
    attendance_present_count = Attendance.objects.filter(
            student=student,
            date__range=[start_date, end_date],
            present=True
        ).count()
    attendance_absent_count = (end_date - start_date).days + 1 - attendance_present_count - holidays_count  #
    

    context = {
        "holiday_days":holiday_days  ,
        "month":month,
        'student': student,
        'attendance_data': attendance_data,
        'weeks': weeks,
        'classgroup': classgroup,
        'today': today,
        "holidays_count": holidays_count,
'attendance_absent_count':attendance_absent_count,
'attendance_present_count':attendance_present_count,
}
    return render(request, 'students/students_attendance_calendar.html', context)

def attendance_data_students(request, pk):
    end_date = datetime.now().date()
    start_date = end_date - timedelta(days=6)  # Last 7 days
    classgroup = ClassGroup.objects.get(id=pk)
    students = Student.objects.filter(classgroup=classgroup)
    holidays = Holiday.objects.filter(date__range=[start_date, end_date])
    attendance_data = {}
    for student in students:
        student_attendance = {}
        for date in (start_date + timedelta(days=i) for i in range(7)):
            holiday = holidays.filter(date=date).first()
            if holiday or date.weekday() == 6:
                student_attendance[date] = f"Holiday " 
            else:
                is_present = Attendance.objects.filter(student=student, date=date,present=True).exists()
                student_attendance[date] = "Present" if is_present else "Absent"
        attendance_data[student.id] = student_attendance

    dates = [start_date + timedelta(days=i) for i in range(7)]  #  7 dates

    context = {
        'students': students,
        'attendance_data': attendance_data,
        'dates': dates,
        'classgroup': classgroup,
    }
    return render(request, 'students/attendance_table.html', context)
@login_required(login_url='login-page')
def student_attendance_list(request, pk):
    holiday= Holiday.objects.filter(date=datetime.today().date()).exists()
    if holiday:
        return redirect('holidays',classid=pk)
    user = request.user
    if isinstance(user, CustomUser):
        username = user.username
        user_type = user.user_type
    classgroup = get_object_or_404(ClassGroup, id=pk)
    students = Student.objects.filter(classgroup=classgroup)
    today = timezone.now().date() 
    student_attendance = []
    for student in students:
        attendance_exists = Attendance.objects.filter(
            Q(student=student) & Q(date=today) & Q(student__classgroup=classgroup.id) & Q(present=True)
        ).exists() 
        student_attendance.append((student, attendance_exists))
    context = {'student_attendance': student_attendance, "today": today, "classgroup": classgroup}
    return render(request, 'students/attendance.html', context)
@login_required(login_url='login-page')
def past_attendance_student_list(request, pk):
    holiday= Holiday.objects.filter(date=datetime.today().date()).exists()
    if holiday:
        return redirect('holidays')
    user = request.user
    if isinstance(user, CustomUser):
        username = user.username
        user_type = user.user_type
    classgroup = get_object_or_404(ClassGroup, id=pk)
    students = Student.objects.filter(classgroup=classgroup)
    today = timezone.now().date() 
    student_attendance = []
    for student in students:
        attendance_exists = Attendance.objects.filter(
            Q(student=student) & Q(date=today) & Q(student__classgroup=classgroup.id) & Q(present=True)
        ).exists() 
        student_attendance.append((student, attendance_exists))
    context = {'student_attendance': student_attendance, "today": today, "classgroup": classgroup}
    return render(request, 'students/attendance.html', context)
def create_subjescts(request):
    forms=subjectform() 
    if request.method == "POST":
        forms=subjectform(request.POST)
        if forms.is_valid():
            forms.save() 
            return redirect('home')
    content={'forms':forms}
    return render(request,"students/create_subjects.html",content)
def teacher_list(request):
    user=request.user 
    username=user.username
    if isinstance(user, CustomUser):    
        customuser = user
        user_type = user.user_type 
    teachers = Teacher.objects.all() 
    today = timezone.now().date() 
    print(user_type,"username",customuser.username)
    content={"today":today ,'customuser':customuser ,
             'user_type':user_type, 'teachers':teachers} 
    return render(request, 'students/teacher_data.html', content)  
    
    
def send_mails_to_teachers(request): 
    subject=request.POST.get('subject')  
    mail_text=request.POST.get('content')  
    if request.method == 'POST': 
        send_mail_to_all_task_teacher.delay(subject,mail_text) 
        return redirect('home') 
    content={} 
    return render(request,"base/send_mail.html",content) 

def send_mails_to_students(request,id): 
    subject=request.POST.get('subject')  
    mail_text=request.POST.get('content')  
    if request.method == 'POST': 
        send_mail_task.delay(id,subject,mail_text) 
        
        return redirect('home') 
    content={} 
    return render(request,"base/send_mail.html",content) 

def send_mails_to_all_students(request ,classid):
     
    subject=request.POST.get('subject')  
    mail_text=request.POST.get('content')  
    if request.method == 'POST': 
        send_mail_to_all_task_students.delay(subject,mail_text,classid) 
        return redirect('home') 
    content={} 
    return render(request,"base/send_mail.html",content)
def check_presenty(id): 
    today=timezone.now().date() 
    student=Student.objects.get(id=id) 
    attendance=Attendance.objects.get_or_create(student= student ,date=today)
    return attendance.present 

def mark_attendance(request, student_id):
    student = get_object_or_404(Student, id=student_id)
    today = timezone.now().date()
    attendance, created = Attendance.objects.get_or_create(student=student, date=today)
    attendance.present = True
    attendance.save()  
    return redirect('home')

def mark_apsent(request, student_id):
    student=get_object_or_404(Student,id=student_id)
    today=timezone.now().date()
    attendance=Attendance.objects.get(student=student,date=today)
    attendance.present=False
    attendance.save()
    return redirect("home") 


def id_card(request,id):
    student=Student.objects.get(id=id) 
    content={'student':student}
    return render(request,'students/id_card.html',content)




def export_data_excel(request, class_id):
    classgroup = ClassGroup.objects.get(id=class_id)

    end_date = datetime.now().date()
    start_date = end_date - timedelta(days=30)
    students = Student.objects.filter(classgroup=classgroup)
    data = []
    dates = [start_date + timedelta(days=i) for i in range(31)]
    holidays=Holiday.objects.filter(date__range=[start_date, end_date])
    holiday_dates = {hol_date.date for hol_date in holidays}
    for student in students:
        row = {
            'Student ID': student.roll_no,
            'Name': student.name,
        }
        
        for date in dates:
            if date in holiday_dates :
                row[date.strftime('%Y-%m-%d')] = 'Holiday(F)'
            elif date.weekday() == 6: 
                row[date.strftime('%Y-%m-%d')] = 'Holiday(S)'
            else:
                attendance = Attendance.objects.filter(student=student, date=date).first()
                row[date.strftime('%Y-%m-%d')] = 'Present' if attendance and attendance.present else 'Absent'
        data.append(row)
    
    df = pd.DataFrame(data)
    
    # Create an Excel writer using pandas with openpyxl engine
    file_path = rf'students\management\{classgroup.name}-{classgroup.year}.xlsx'
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name=f'{classgroup.name}-{classgroup.year} Attendance')
    
    # Load the workbook and select the active sheet
    workbook = load_workbook(file_path)
    sheet = workbook.active
    
    # Adjust column widths
    for column in sheet.columns:
        max_length = 0
        column_name = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column_name].width = adjusted_width
    
    workbook.save(file_path)
    
    if os.path.exists(file_path):
        with open(file_path, 'rb') as file:
            response = HttpResponse(file.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename={os.path.basename(file_path)}'
            return response
    else:
        raise Http404("File not found")
def attendance_export_date_form(request,class_id):
    classgroup = ClassGroup.objects.get(id=class_id)
    if request.method == 'POST':
        start_date = datetime.strptime(request.POST['start_date'], '%Y-%m-%d').date()
        end_date = datetime.strptime(request.POST['end_date'], '%Y-%m-%d').date()
        return redirect('export_data_exel_past', class_id=class_id, start_date=start_date.strftime('%Y-%m-%d'), end_date=end_date.strftime('%Y-%m-%d'))
    else:
        error(request,"please select a date")
    return render(request,'students/attendance_export_date_form.html')
    
def export_past_data_excel(request, class_id, start_date, end_date):
    classgroup = ClassGroup.objects.get(id=class_id)
    students = Student.objects.filter(classgroup=classgroup)
    start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    delta = (end_date - start_date).days + 1
    dates = [start_date + timedelta(days=i) for i in range(delta)]
    holidays = Holiday.objects.filter(date__range=[start_date, end_date])
    holiday_dates = {hol.date for hol in holidays}

    data = []
    for student in students:
        row = {
            'Student ID': student.roll_no,
            'Name': student.name,
        }
        
        for date in dates:
            if date in holiday_dates:
                row[date.strftime('%Y-%m-%d')] = 'Holiday(F)'
            elif date.weekday() == 6:  # Sunday check
                row[date.strftime('%Y-%m-%d')] = 'Holiday(S)'
            else:
                attendance = Attendance.objects.filter(student=student, date=date).first()
                row[date.strftime('%Y-%m-%d')] = 'Present' if attendance and attendance.present else 'Absent'
        
        data.append(row)
    
    # Create DataFrame and export to Excel
    df = pd.DataFrame(data)
    file_path = os.path.join('students', 'management', f'{classgroup.name}-{classgroup.year}.xlsx')
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name=f'{classgroup.name}-{classgroup.year} Attendance')
    
    # Adjust column widths
    workbook = load_workbook(file_path)
    sheet = workbook.active
    for column in sheet.columns:
        max_length = 0
        column_name = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column_name].width = adjusted_width
    workbook.save(file_path)
    
    # Serve the file as a downloadable response
    if os.path.exists(file_path):
        with open(file_path, 'rb') as file:
            response = HttpResponse(file.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename={os.path.basename(file_path)}'
            return response
    else:
        raise Http404("File not found")
def import_data(request,classid):
    Fake=Faker()
    classes = Classes.objects.all()
    class_ba=Classes.objects.get(name="BA")    
    years = ClassGroup.objects.filter(name=class_ba)
    if request.method == 'POST':
        try:
            file_path = request.FILES.get('export_data')
            df = pd.read_excel(file_path)
            class_obj=request.POST.get('class')
            year=request.POST.get('year')

            classgroup=ClassGroup.objects.get(name=class_obj,year=year)   
            for _, row in df.iterrows(): 
                Student.objects.create(
                    roll_no=row['roll no'],
                    name=row['Name'],
                    contact=Fake.phone_number() ,
                    classgroup=classgroup,
                )
            return redirect('attendance-list',pk=classid)
            # stdout.write(style.SUCCESS('Successfully imported student data'))

        except Exception as e:
            error(request,f"Error:  {e}")
    content={"classes":classes,"years":years}
    return render(request,"students/class_selecting.html",content)


def create_holiday(request):
    form=holidays_form()
    if request.method == "POST":
        form=holidays_form(request.POST)
        if form.is_valid():
            form.save()
            return redirect("all-classes")
        else:
            error(request,"Please enter valid Data")
    content={"form":form}
    return render(request,"students/create_holiday.html",content)
def Holiday_page(request,classid):
    classgroup=ClassGroup.objects.get(id=classid)
    holiday = Holiday.objects.get(date=datetime.today().date())
    content={"holiday":holiday,
             "classgroup":classgroup}
    return render(request,"students/Holidays.html",content)





import os
import cv2
import numpy as np
import torch
import base64
import pygame
import threading
import time

from datetime import datetime
from django.conf import settings
from django.shortcuts import render, redirect
from django.core.files.base import ContentFile
from django.utils import timezone
from django.contrib import messages
from django.contrib.auth.decorators import login_required

from facenet_pytorch import MTCNN, InceptionResnetV1

from .models import Student, Attendance, Subject, CameraConfiguration


# Initialize face detection and recognition models
mtcnn = MTCNN(keep_all=True)
resnet = InceptionResnetV1(pretrained='vggface2').eval()



def detect_and_encode(image):
    """Detect and encode faces in an image."""
    with torch.no_grad():
        boxes, _ = mtcnn.detect(image)
        if boxes is not None:
            faces = []
            for box in boxes:
                face = image[int(box[1]):int(box[3]), int(box[0]):int(box[2])]
                if face.size == 0:
                    continue
                face = cv2.resize(face, (160, 160))
                face = np.transpose(face, (2, 0, 1)).astype(np.float32) / 255.0
                face_tensor = torch.tensor(face).unsqueeze(0)
                encoding = resnet(face_tensor).detach().numpy().flatten()
                faces.append(encoding)
            return faces
    return []

def encode_uploaded_images():
    """Load known student face encodings."""
    encodings = []
    names = []
    students = Student.objects.all()
    for student in students:
        img_path = os.path.join(settings.MEDIA_ROOT, str(student.profile))
        img = cv2.imread(img_path)
        if img is None:
            continue
        rgb_img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
        face_enc = detect_and_encode(rgb_img)
        if face_enc:
            encodings.extend(face_enc)
            names.append(student.name)
    return encodings, names

def recognize_faces(known_encodings, known_names, test_encodings, threshold=0.6):
    """Compare test encodings with known ones and return recognized names."""
    recognized = []
    for test_enc in test_encodings:
        distances = np.linalg.norm(known_encodings - test_enc, axis=1)
        min_index = np.argmin(distances)
        if distances[min_index] < threshold:
            recognized.append(known_names[min_index])
        else:
            recognized.append("Not Recognized")
    return recognized

def capture_student(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        email = request.POST.get('email')
        phone = request.POST.get('phone_number')
        student_class = request.POST.get('student_class')
        image_data = request.POST.get('image_data')

        if image_data:
            header, encoded = image_data.split(',', 1)
            image_file = ContentFile(base64.b64decode(encoded), name=f"{name}.jpg")

            Student.objects.create(
                name=name,
                email=email,
                phone_number=phone,
                student_class=student_class,
            )
            return redirect('selfie_success')
    return render(request, 'capture_student.html')


def selfie_success(request):
    return render(request, 'selfie_success.html')


# ---------- FACE RECOGNITION AND ATTENDANCE ----------

def capture_and_recognize(request):
    stop_events = []
    threads = []
    windows = []
    errors = []

    subject_id = request.GET.get('subject_id')  # Example: ?subject_id=1
    subject = None
    if subject_id:
        subject = Subject.objects.filter(id=subject_id).first()

    def process_camera(cam_config, stop_event):
        cap = None
        try:
            source = cam_config.camera_source.strip()
            if source.isdigit():
                cap = cv2.VideoCapture(1)
            else:
                cap = cv2.VideoCapture(source + "/video" if not source.endswith("/video") else source)

            if not cap.isOpened():
                raise Exception(f"Cannot open camera: {cam_config.name}")


            win_name = f"Recognition - {cam_config.name}"
            windows.append(win_name)

            known_encs, known_names = encode_uploaded_images()

            while not stop_event.is_set():
                ret, frame = cap.read()
                if not ret:
                    break
                rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                test_encs = detect_and_encode(rgb)

                if test_encs and known_encs:
                    results = recognize_faces(np.array(known_encs), known_names, test_encs, cam_config.threshold)

                    for name, box in zip(results, mtcnn.detect(rgb)[0]):
                        if box is not None:
                            (x1, y1, x2, y2) = map(int, box)
                            cv2.rectangle(frame, (x1, y1), (x2, y2), (0, 255, 0), 2)
                            cv2.putText(frame, name, (x1, y1 - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.9, (255, 255, 255), 2)

                            if name != "Not Recognized":
                                student = Student.objects.filter(name=name).first()
                                if student:
                                    today = datetime.now().date()
                                    existing = Attendance.objects.filter(student=student, date=today, subject=subject).first()
                                    if not existing:
                                        Attendance.objects.create(
                                            student=student,
                                            date=today,
                                            present=True,
                                            subject=subject
                                        )
                                        cv2.putText(frame, f"{name} marked present", (50, 50), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 255, 0), 2)

                cv2.imshow(win_name, frame)
                if cv2.waitKey(1) & 0xFF == ord('q'):
                    stop_event.set()
                    break
        except Exception as e:
            errors.append(f"{cam_config.name}: {str(e)}")
        finally:
            if cap:
                cap.release()
            cv2.destroyWindow(win_name)
    try:
        cam_configs = CameraConfiguration.objects.all()
        if not cam_configs.exists():
            raise Exception("No cameras configured in admin panel.")
        for cam_config in cam_configs:
            stop_event = threading.Event()
            stop_events.append(stop_event)
            t = threading.Thread(target=process_camera, args=(cam_config, stop_event))
            threads.append(t)
            t.start()
        while any(t.is_alive() for t in threads):
            time.sleep(1)

    except Exception as e:
        errors.append(str(e))

    finally:
        for event in stop_events:
            event.set()

        for win in windows:
            if cv2.getWindowProperty(win, cv2.WND_PROP_VISIBLE) >= 1:
                cv2.destroyWindow(win)

    if errors:
        return render(request, 'base/error.html', {'error_message': '\n'.join(errors)})

    return redirect('home')


